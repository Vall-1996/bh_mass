# -*- coding: utf-8 -*-
"""
Created on Fri Jun 28 13:52:17 2019

@author: div
"""

def make_database(folder_path,date_file,loc_file,ranges):
    """
    Build the graph set [] with subsets date+times->sources, region->
    coords+source, source->dates+times, source->region+coords, date+times->
    region->coords.
    """  
    from openpyxl import load_workbook
    from collections import defaultdict
    
    wbd = load_workbook(folder_path+date_file)
    sheet_dates = wbd.active
    ft = defaultdict(list) #filename = [(datetime_start,datetime_end),...]
    id_name,date=-1,-1
    for i in range(ranges[0][0],ranges[0][1]):
        if sheet_dates['A'+str(i)].value:
            id_name = str(sheet_dates['A'+str(i)].value)+str(sheet_dates['B'+str(i)].value)
        if sheet_dates['C'+str(i)].value:
            date = str(sheet_dates['C'+str(i)].value)+":"
        time_start = date+str(sheet_dates['D'+str(i)].value)
        time_end = date+str(sheet_dates['E'+str(i)].value)
        ft[id_name].append((time_start,time_end))
        
    wbl = load_workbook(folder_path+loc_file)
    sheet_locs = wbl.active
    fc = defaultdict(list) #filename = (ra,dec)
    for i in range(ranges[1][0],ranges[1][1]):        
        #print(sheet_locs['D'+str(i)].value)
        id_name = str(sheet_locs['A'+str(i)].value)+str(sheet_locs['B'+str(i)].value)
        ra_dec = (sheet_locs['C'+str(i)].value,sheet_locs['D'+str(i)].value)
        #print(ra_dec)
        fc[id_name].append(ra_dec)

    tf = defaultdict(list) #(datetime_start,datetime_end) = [filename,...]
    for key, value in ft.items():
        for v in value:
            tf[v].append(key)
        
    tc = defaultdict(list) #(datetime_start,datetime_end) = [(ra,dec),...]
    for key in tf:
        for value in tf[key]:
            tc[key].append(fc[value])
    
    return [ft,fc,tf,tc]

def flatten_database(db):
    new_ft = {}
    new_tc = {}
    return [new_ft,db[1],new_tc]

def same_obs_period():
    #if within the same night, return True, else False 
    return False



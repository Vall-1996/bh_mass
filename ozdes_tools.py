# -*- coding: utf-8 -*-
"""
Created on Fri Jun  7 16:03:13 2019
Tools for OzDES spectral fits files.
@author: val
"""
def get_spectra(folder_path,file_name,i=0,m=[],mask_break=False,lines=True,plot=False,save=False):
    """
    Obtain spectra as wavelength[] and flux[] from an OzDES spectral fits file. 
    Index refers to the ith entry in an extended fits file. Produce a plot 
    onscreen and/or save a plot via the plot/save flags. Mask the stitching
    break seen in some spectra by setting mask_break flag to True. Underlay
    spectral lines to the plot by setting lines flag to True. Pass in m values 
    to perform a gaussian kernel convolution with additional smoothing. Note
    that the spectra is redshifted and the flux calibrated. For coadds, the
    exposure time is assumed the average exposure of the coadded files.
    """
    from astropy.io import fits
    import matplotlib.pyplot as plt
    import numpy as np
    a = fits.open(folder_path+file_name)
    if i < len(a) and i >= 0:
        flux=a[i].data
        dispersion_start = a[i].header['CRVAL1'] - (a[i].header['CRPIX1'] - 1) * a[i].header['CDELT1']
        wavelength = dispersion_start + a[i].header['CDELT1']*np.arange(0, len(flux))
        if mask_break:
            wavelength,flux,temp=remove_range(wavelength,flux,[5550,5600])
        flux=exposure(wavelength,flux,a,i)
        wavelength = redshift(wavelength,a[0].header['Z']) 
        #if b > 0:
        #    from astropy.convolution import convolve, Box1DKernel
        #    flux = convolve(flux, Box1DKernel(b))
        #    wavelength = wavelength[b+1:-b-1]
        #    flux = flux[b+1:-b-1]
        if len(m) == 2:
            from astropy.convolution import Gaussian1DKernel, convolve
            import scipy.signal
            def smooth(y,std): #(std should be 1 or 2 pixels)
                gaussian_smooth = Gaussian1DKernel(stddev=std)
                y_smooth=convolve(y,gaussian_smooth)
                return y_smooth
            flux = smooth(flux,m[0]) 
            flux = scipy.signal.medfilt(flux,m[1])#m[1] should be odd
        if plot:
            
            #wavelength=wavelength[300:]
            #flux=flux[300:]
            wlrange=[5550/(1+a[0].header['Z']),5600/(1+a[0].header['Z'])]
            wavelength,flux,tt=remove_range(wavelength,flux,wlrange)
            
            plt.figure(figsize=(10, 4))
            plt.title(file_name)
            plt.scatter(wavelength,flux,s=2,c='black',marker='.')
            #plt.plot(wavelength, flux, 'k-')
            if lines:
                plt.axvspan(5550/(1+a[0].header['Z']),5600/(1+a[0].header['Z']),color='gray',alpha=0.1)
                imp_lines=[2799.117,4862.68,6564.61]
                #imp_lines=[2799.117,4862.68,4341.68,4102.89,6564.61]
                for x in imp_lines:
                    if x>wavelength[0] and x<wavelength[-1]:
                        plt.axvline(x=x,color='red',alpha=0.3)                
                other_lines=[1908.734,2326.0,3727.092,4960.295,5008.240,6549.86,6585.27,6718.29,6732.67]
                #for x in other_lines:
                #    if x>wavelength[0] and x<wavelength[-1]:
                #        plt.axvline(x=x,color='blue',alpha=0.3)  
                #abs_lines=[3934.777,3969.588,4305.61,5176.7,5895.6,8500.36,8544.44,8664.52]
                #for x in abs_lines:
                #    if x>wavelength[0] and x<wavelength[-1]:
                #        plt.axvline(x=x,color='green',alpha=0.3)
                #wgtls_lines=[2439.5,3346.79,3426.85,3729.875,3889.0,4072.3,4364.436,4932.603,6302.046,6365.536,6529.03]
                #for x in wgtls_lines:
                #    if x>wavelength[0] and x<wavelength[-1]:
                #        plt.axvline(x=x,color='cyan',alpha=0.1)
                #sky_lines=[5578.5/(1+a[0].header['Z']),5894.6/(1+a[0].header['Z']),6301.7/(1+a[0].header['Z'])]
                #for x in sky_lines:
                #    if x>wavelength[0] and x<wavelength[-1]:
                #        plt.axvline(x=x,color='magenta',alpha=0.3)
            plt.xlabel("Wavelength [A]")
            plt.ylabel("Flux [10^-23 erg/s/cm^2/A]")
            plt.show()
        if save:
            plt.figure(figsize=(16, 6))
            plt.title(file_name)
            plt.plot(wavelength, flux, 'k-')
            if lines:
                plt.axvspan(5550/(1+a[0].header['Z']),5600/(1+a[0].header['Z']),color='gray',alpha=0.1)
                imp_lines=[2799.117,4862.68,6564.61]
                for x in imp_lines:
                    if x>wavelength[0] and x<wavelength[-1]:
                        plt.axvline(x=x,color='red',alpha=0.6)                
                other_lines=[1908.734,2326.0,3727.092,4102.89,4341.68,4960.295,5008.240,6549.86,6585.27,6718.29,6732.67]
                for x in other_lines:
                    if x>wavelength[0] and x<wavelength[-1]:
                        plt.axvline(x=x,color='blue',alpha=0.3)
                abs_lines=[3934.777,3969.588,4305.61,5176.7,5895.6,8500.36,8544.44,8664.52]
                for x in abs_lines:
                    if x>wavelength[0] and x<wavelength[-1]:
                        plt.axvline(x=x,color='green',alpha=0.3)
                wgtls_lines=[2439.5,3346.79,3426.85,3729.875,3889.0,4072.3,4364.436,4932.603,6302.046,6365.536,6529.03]
                for x in wgtls_lines:
                    if x>wavelength[0] and x<wavelength[-1]:
                        plt.axvline(x=x,color='cyan',alpha=0.1)
                #sky_lines=[5578.5/(1+a[0].header['Z']),5894.6/(1+a[0].header['Z']),6301.7/(1+a[0].header['Z'])]
                #for x in sky_lines:
                #    if x>wavelength[0] and x<wavelength[-1]:
                #        plt.axvline(x=x,color='magenta',alpha=0.3)
            plt.xlabel("Wavelength [A]")
            plt.ylabel("Flux [10^-23 erg/s/cm^2/A]")
            plt.show()
            plt.savefig(folder_path+"fig_"+file_name[:-5]+"_"+str(i)+".png")
            plt.close()
        return wavelength,flux,a[0].header['Z']
    else:
        print("Fits file index out of range. Range is 0 to "+str(len(a)-1)+".")
        return -1,-1  

def plot_all_spectra(folder_path,file_name,b=0,m=[],mask_break=False,save=False,show_coadd=False):
    """
    Plot all the spectra in a fits file overlayed. Save a plot via the save 
    flag. Mask the stitching break seen in some spectra by setting mask_break 
    flag to True. Overlay the coadd spectra by setting show_coadd flag to True.
    """
    from astropy.io import fits
    import matplotlib.pyplot as plt
    a = fits.open(folder_path+file_name)
    plt.figure()
    for index in range(3,len(a),3):
        wavelength, flux, z = get_spectra(folder_path,file_name,i=index,m=m,mask_break=mask_break)
        plt.plot(wavelength, flux, 'k-',alpha=0.4)
    if show_coadd:
        wavelength, flux, z = get_spectra(folder_path,file_name,i=0,m=m,mask_break=mask_break)
        plt.plot(wavelength, flux, 'r-',alpha=0.6)
    plt.show()
    if save:
        plt.savefig(folder_path+"fig_"+file_name[:-5]+"_all.png")
        plt.close()
    
def get_range(wavelength,flux,wavelength_range):
    """
    Obtain a wavelength range within a larger spectral data set.
    """
    if wavelength_range[0]>=wavelength_range[1]:
        print("Invalid wavelength range. Initial must be smaller than final.")
        return None,None
    index_min,index_max = -1,-1
    for i in range(len(wavelength)-1):
        if wavelength[i] <= wavelength_range[0] and wavelength[i+1] > wavelength_range[0]:
            index_min=i
        if wavelength[i] <= wavelength_range[1] and wavelength[i+1] > wavelength_range[1]:
            index_max=i
            break
    if index_min < 0 or index_max < 0:
        print("Wavelength range is out of bounds. Bounds are "+str(int(wavelength[0])+1)+" to "+str(int(wavelength[len(wavelength)-1])-1)+".")
        return None, None
    xdata = wavelength[index_min:index_max]
    ydata = flux[index_min:index_max]
    return xdata, ydata

def remove_range(wavelength,flux,wavelength_range):
    """
    Remove a wavelength range within a larger spectral data set.
    """
    import numpy as np
    if wavelength_range[0]>=wavelength_range[1]:
        print("Invalid wavelength range. Initial must be smaller than final.")
        return -1
    index_min,index_max = -1,-1
    for i in range(len(wavelength)-1):
        if wavelength[i] <= wavelength_range[0] and wavelength[i+1] > wavelength_range[0]:
            index_min=i
        if wavelength[i] <= wavelength_range[1] and wavelength[i+1] > wavelength_range[1]:
            index_max=i
            break
    if index_min < 0 or index_max < 0:
        print("Wavelength range is out of bounds. Bounds are "+str(int(wavelength[0])+1)+" to "+str(int(wavelength[len(wavelength)-1])-1)+".")
    temp_1 = wavelength[0:index_min]
    temp_2 = wavelength[index_max:len(wavelength)-1]
    xdata = np.hstack([temp_1,temp_2])
    temp_1 = flux[0:index_min]
    temp_2 = flux[index_max:len(flux)-1]
    ydata = np.hstack([temp_1,temp_2])
    return xdata, ydata, [index_min,index_max]

def redshift(wavelength,z):
    """
    Calculate a redshifted wavelength where redshift = z. Used to redshift
    spectra into the rest frame.
    """
    wavelength_z=[]
    for x in wavelength:
        wavelength_z.append(x/(1.0+z))
    return wavelength_z

def redshift_effect(z,x,dz=None,dx=None):
    """
    Find the change in wavelength (dx) or in redshift (dz) given a wavelength
    and a redshift with some potential error (dx or dz).
    """
    result = -1
    if dx is not None and dz is None:
        result = x/((x/(1+z))+dx)-(1+z)
    if dz is not None and dx is None:
        result = x/(1+z+dz)-x/(1+z)
    return result

def adjust_spectra(wavelength,flux,z,dz):
    """
    Given a redshift with some error, adjust the spectrum wavelength and flux.
    """
    wavelength_dz,flux_dz=[],[]
    for i in range(len(wavelength)):
        x_dx=wavelength[i]+redshift_effect(z,wavelength[i],dz=dz)
        wavelength_dz.append(x_dx)
        flux_dz.append(flux[i]*wavelength[i]/x_dx)
    return wavelength_dz,flux_dz

def exposure(wavelength,flux,a,i):
    """
    Given a spectrum wavelength and flux, calibrate the flux using either the
    exact exposure or an average exposure (for coadds).
    """
    if i%3 == 0:
        if i == 0:
            import statistics as stats
            exp=[]
            for j in range(3,len(a),3):
                exp.append(a[j].header['EXPOSED'])
            exp=stats.mean(exp)
            #print("Correcting for exposure using average of "+str(exp))
        else:
            exp = float(a[i].header['EXPOSED'])
            #print("Correcting for exposure using header of "+str(exp))
        flux_e=[]
        constant = (6.626*2.998)*(10**8)/(225*exp)
        for i in range(len(wavelength)):
            flux_e.append(flux[i]*constant/wavelength[i])
        print("Please note the fluxes have been scaled by a factor of 10^23 erg/s/cm^2/A.")
        return flux_e
    else:
        return flux

def calculate_mass(values,errors,verbose=False):
    """
    Given values with errors of the peak center, the peak width, the integral
    difference between the peak and the continuum, the redshift, the c and d
    fit parameters (given the emission line) and the luminosity distance,
    calculate the mass of the black hole. If the luminosity distance is 
    unknown, a lookup calculator is called by setting the luminosity distance
    to -1. This increases runtime drastically, so the result is printed and
    can be used for future calculations at the given redshift. Black hole
    mass result m is given in the format of 10^M solar masses, M is returned.
    """
    #center,width,intEC,z,c,d,DL
    import numpy as np
    light = 299792 #km/s
    #msun = 1.98847*(10**30) #kg
    if values[6]<0: #in cm, flat
        values[6]=lookup_DL(values[3]) 
    values[6]=values[6]*3.08568#*(10**24)
    FWHM = light*(values[1]/values[0])
    Lline = ((4*np.pi*(values[6]**2))*values[2])/(10**19)
    if verbose:
        print('FWHM:  '+str("{:.2f}".format(FWHM)))
        print('Lline: '+str("{:.10f}".format(Lline)))
    dx = abs(values[5]*np.log10(Lline))*np.sqrt((errors[5]/values[5])**2+(0.434*errors[2]/(values[2]))**2)
    dy = 0.868*np.sqrt((errors[0]/values[0])**2+(errors[1]/values[1])**2)#light taken out..
    return values[4]+values[5]*np.log10(Lline)+2*np.log10(FWHM),np.sqrt(errors[4]**2+dx**2+dy**2)

def lookup_DL(z):
    """
    Look up a luminosity distance given a redshift of z. Requires internet.
    """
    from selenium import webdriver
    print("Looking up DL from z online...")
    print("Accessed http://www.astro.ucla.edu/~wright/CosmoCalc.html")
    print("Inputting a redshift of "+str(z))
    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    options.add_argument('window-size=1366x768')
    options.add_argument("disable-gpu")
    driver = webdriver.Chrome(chrome_options=options)
    driver.get('http://www.astro.ucla.edu/~wright/CosmoCalc.html')
    driver.switch_to.frame('CCform')
    redshift_box = driver.find_element_by_name('txtz')
    redshift_box.clear()
    redshift_box.send_keys(str(z))
    flat_button = driver.find_element_by_name('Flat')
    flat_button.click()
    general_button = driver.find_element_by_name('General')
    general_button.click()
    driver.switch_to_default_content()
    driver.switch_to.frame('CCout')
    text = driver.find_element_by_xpath('/html/body/ul/li[8]').text
    print("DL calculated to be "+text[30:-18]+" Mpc.")
    return float(text[30:-18])

def integrate(gfit,lfit,wavelength_range,plot_values=False):
    """
    Given a linear fit (lfit) and a gaussian fit (gfit), find the integral
    difference between the wavelength range given. If plot_values flag is True,
    also return data which can be used for plotting purposes. Results are 
    returned as gaussian integral with error, linear integral with error.
    """
    from scipy.integrate import quad
    import numpy as np
    g=gfit[2]/np.sqrt(2*np.log(2))
    f=gfit[0]/(g*np.sqrt(2*np.pi))
    gauss = lambda x: gfit[3]+f*np.exp(-((x-gfit[1])**2)/(2*(g**2)))
    line = lambda x: lfit[0]*x+lfit[1]
    if plot_values:
        def gaussian(x,A,u,o,b):  
            g = o/np.sqrt(2*np.log(2))
            gauss = b+(A/(g*np.sqrt(2*np.pi)))*np.exp(-((x-u)**2)/(2*(g**2)))
            return gauss
        xdata,ldata,gdata=[],[],[]
        for i in range(int(wavelength_range[0]-2),int(wavelength_range[1]+2)):
            xdata.append(i)
            gdata.append(gaussian(i,gfit[0],gfit[1],gfit[2],gfit[3]))
            ldata.append(lfit[0]*i+lfit[1])
        return quad(gauss, wavelength_range[0], wavelength_range[1]), quad(line, wavelength_range[0], wavelength_range[1]),[xdata,gdata,ldata]
    return quad(gauss, wavelength_range[0], wavelength_range[1]), quad(line, wavelength_range[0], wavelength_range[1])

def find_intersects(lfit,gfit,wavelength_range,step):
    """
    Given a peak gaussian fit and a continuum linear fit, find the two nearest
    points within a wavelength range by combing over the range with a step
    size. This is used to find an appropriate range to integrate over later.
    """
    import numpy as np
    #Defining the continuum and peak functions.
    def gaussian(x,A,u,o,b):  
        g = o/np.sqrt(2*np.log(2))
        gauss = b+(A/(g*np.sqrt(2*np.pi)))*np.exp(-((x-u)**2)/(2*(g**2)))
        return gauss
    def line(x,a,b):
        return a*x+b
    #Finding the nearest pair to the left of the peak center.
    x = wavelength_range[0]
    closeness=1000000
    intersect1 = -1
    while(x<=gfit[1]):
        peak = gaussian(x,gfit[0],gfit[1],gfit[2],gfit[3])
        cont = line(x,lfit[0],lfit[1])
        if abs(peak-cont) < closeness:
            #Update the nearest intersect found.
            closeness = abs(peak-cont)
            intersect1=x
        x=x+step
    #Finding the nearest pair to the right of the peak.
    x=gfit[1]
    closeness=1000000
    intersect2 = -1
    while(x<=wavelength_range[1]):
        peak = gaussian(x,gfit[0],gfit[1],gfit[2],gfit[3])
        cont = line(x,lfit[0],lfit[1])
        if abs(peak-cont) < closeness:
            #Update the nearest intersect found.
            closeness = abs(peak-cont)
            intersect2=x
        x=x+step
    return [intersect1,intersect2]

def fit_line(wavelength,flux,initial_fit=[1.0,1.0],wavelength_range=[],plot=False,verbose=False):
    """
    Fit a line to the continuum described by wavelength and flux between some
    wavelength range with an initial fit of [a,b]. Plot the fit result if plot
    flag is set to True. Print the fit result if verbose flag is set to True.
    """
    from scipy import optimize as opt
    import matplotlib.pyplot as plt
    import numpy as np
    def line(x,a,b):
        return a*x+b
    if len(wavelength_range) == 2:
        wavelength,flux=get_range(wavelength,flux,wavelength_range)
    popt, pcov = opt.curve_fit(line, wavelength, flux, initial_fit)
    perr = np.sqrt(np.diag(pcov))
    if verbose:
        print('a: '+str("{:.7f}".format(popt[0]))+' +/- '+str("{:.7f}".format(perr[0])))
        print('b: '+str("{:.7f}".format(popt[1]))+' +/- '+str("{:.7f}".format(perr[1])))
    fdata = []
    for x in wavelength:
        fdata.append(line(x,popt[0],popt[1]))
    if plot:
        plt.figure()
        plt.plot(wavelength, flux, 'k-')
        plt.plot(wavelength, fdata, 'r-')
        plt.show()
    return wavelength,fdata,[popt,perr]

def fit_gaussian(wavelength,flux,initial_fit=[1.0,1.0,1.0,1.0],wavelength_range=[],plot=False,verbose=False):
    """
    Fit a gaussian to the peak described by wavelength and flux between some
    wavelength range with an initial fit of [A,o,u,b]. Plot the fit result if 
    plot flag set to True. Print the fit result if verbose flag set to True.
    """
    from scipy import optimize as opt
    import matplotlib.pyplot as plt
    import numpy as np
    def gaussian(x,A,u,o,b):
        import numpy as np    
        g = o/np.sqrt(2*np.log(2))
        gauss = b+(A/(g*np.sqrt(2*np.pi)))*np.exp(-((x-u)**2)/(2*(g**2)))
        return gauss
    if len(wavelength_range) == 2:
        wavelength,flux=get_range(wavelength,flux,wavelength_range)
    bounds=((0.0,0.0,0.0,-np.inf),(np.inf,np.inf,np.inf,np.inf))
    popt, pcov = opt.curve_fit(gaussian, wavelength, flux, initial_fit, bounds=bounds)
    perr = np.sqrt(np.diag(pcov))
    if verbose:
        print('A: '+str("{:.7f}".format(popt[0]))+' +/- '+str("{:.7f}".format(perr[0])))
        print('u: '+str("{:.7f}".format(popt[1]))+' +/- '+str("{:.7f}".format(perr[1])))
        print('o: '+str("{:.7f}".format(popt[2]))+' +/- '+str("{:.7f}".format(perr[2])))
        print('b: '+str("{:.7f}".format(popt[3]))+' +/- '+str("{:.7f}".format(perr[3])))
    fdata = []
    for x in wavelength:
        fdata.append(gaussian(x,popt[0],popt[1],popt[2],popt[3]))
    if plot:
        plt.figure()
        plt.plot(wavelength, flux, 'k-')
        plt.plot(wavelength, fdata, 'r-')
        plt.show()
    return wavelength,fdata,[popt,perr]

def autorun_fit(folder_path,filename,initial_fit,m,ks,intersect_limit,plots,zoom,DL,c,d):
    """
    For my convenience. Please do not use this as the specifics may not apply.
    """    
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches    
    
    wavelength,flux,z=get_spectra(folder_path, filename, i=0, m=m, mask_break=True,plot=plots[0])
    wavelength_range=[initial_fit[1]-ks[0],initial_fit[1]+ks[1]]
    old_x,old_y=get_range(wavelength,flux,[wavelength_range[0]-ks[2],wavelength_range[1]+ks[3]])
    xdata_peak,fdata_peak,gfit=fit_gaussian(wavelength,flux,initial_fit,wavelength_range,verbose=True)
    new_w,new_f,ind=remove_range(old_x,old_y,wavelength_range)
    xdata_line,fdata_line,lfit = fit_line(new_w,new_f)
    intersects = find_intersects(lfit[0],gfit[0],intersect_limit,0.01)
    
    if plots[1]:
        plt.figure()
        if zoom:
            plt.plot(old_x,old_y,'k-')
        else:
            plt.plot(wavelength,flux,'k-')    
        plt.plot(xdata_line,fdata_line,'b-')
        plt.plot(xdata_peak,fdata_peak,'r-')
        plt.xlabel("Wavelength [A]")
        plt.ylabel("Flux [10^-23 erg/s/cm^2/A]")
        black_patch = mpatches.Patch(color='black', label='Spectrum')
        red_patch = mpatches.Patch(color='red', label='Peak')
        blue_patch = mpatches.Patch(color='blue', label='Continuum')
        plt.legend(handles=[black_patch,blue_patch,red_patch])
        plt.show()
    
    intE,intC,plot_values=integrate(gfit[0],lfit[0],intersects,plot_values=True)
    print([intE[0]-intC[0],intE[1]+intC[1]])
    
    if plots[2]:
        plt.figure()
        plt.plot(old_x,old_y,'k-')
        plt.plot(plot_values[0],plot_values[1],'r-')
        plt.plot(plot_values[0],plot_values[2],'b-')
        plt.xlabel("Wavelength [A]")
        plt.ylabel("Flux [10^-23 erg/s/cm^2/A]")
        plt.show()
    
    values=[gfit[0][1],gfit[0][2],(intE[0]-intC[0]),z,c[0],d[0],DL]
    errors=[gfit[1][1],gfit[1][2],intE[1]+intC[1],None,c[1],d[1],None]
    mass,error = calculate_mass(values,errors,verbose=True)
    print("Mass: "+"{:.2f}".format(mass)+" ± "+"{:.2f}".format(error))

def irsa_search(folder_path,input_file,output_file,catalog,radius,row_range,input_cols,output_cols):
    """
    Perform an IRSA catalog search on an excel spreadsheet with RA-DEC coords
    given in degrees for the closest matches distance-wise within the catalog,
    and output the name, distance, RA-DEC coords and catalog into specified
    columns. Note: convert HMS and DMS coords into degrees beforehand.
    
    """
    from astroquery.irsa import Irsa
    from astropy.coordinates import SkyCoord
    import astropy.units as u
    from openpyxl import load_workbook
    import numpy as np
    wb = load_workbook(folder_path+input_file)
    sheet = wb.active
    for i in range(row_range[0],row_range[1]):
        ra = sheet[input_cols[0]+str(i)].value
        dec = sheet[input_cols[1]+str(i)].value
        Q = Irsa.query_region(SkyCoord(ra=ra,dec=dec,unit=(u.deg,u.deg)),catalog=catalog,radius=radius*u.arcmin,selcols="object,ra,dec").as_array()   
        if len(Q) < 1:
            continue
        sources_cat = []
        for source in Q:
            dist_cat = np.sqrt((ra-source[1])**2+(dec-source[2])**2)
            sources_cat.append([dist_cat, source[1],source[2],catalog,str(source[0])[2:-1]])
        sorted_sources_cat = sorted(sources_cat, key=lambda x: x[0])
        sheet[output_cols[0]+str(i)]=sorted_sources_cat[0][4]
        sheet[output_cols[1]+str(i)]=sorted_sources_cat[0][0]
        sheet[output_cols[2]+str(i)]=sorted_sources_cat[0][1]
        sheet[output_cols[3]+str(i)]=sorted_sources_cat[0][2]
        sheet[output_cols[4]+str(i)]=sorted_sources_cat[0][3]
    wb.save(folder_path+output_file)    
    return

def plot_emissions(folder_path,filename,wavelength,flux,z,xs,dxs,dzs):
    """
    Diagnostics for finding emission peaks. Given wavelength, flux, redshift
    and emission lines (xs), plot around the emission lines by some value dxs
    and overlay redshift effects at values of dzs.
    """
    import matplotlib.pyplot as plt
    lines=[]
    for i in range(len(xs)):
        w,f=get_range(wavelength,flux,[xs[i]-dxs[i],xs[i]+dxs[i]])
        lines.append([w,f,xs[i],dxs[i]])
    for line in lines:
        if line[0] is not None:
            plt.figure()
            plt.plot(line[0],line[1],'k-')
            plt.axvline(x=line[2],color='b')
            step=0.8/len(dzs)
            a=1.0
            for dz in dzs:
                dx=redshift_effect(x=line[2],z=z,dz=dz)
                if abs(dx)<line[3]:
                    plt.axvline(x=line[2]+dx,color='r',alpha=a-step)
                    plt.axvline(x=line[2]-dx,color='r',alpha=a-step)
                a=a-step
            plt.xlabel("Wavelength [A]")
            plt.ylabel("Flux [10^-23 erg/s/cm^2/A]")
            plt.show()
            plt.savefig(folder_path+"fig_"+filename[:-5]+"_"+str(line[2])+".png")
            plt.close()

def swire_fields_info():
    """
    Remind the user about the SWIRE field names, coordinates and catalogs.
    """    
    print("___NAME___\t___RA___\t___DEC___\t____IRSA_CAT____")
    print("ELAIS-S1\t  9.6250\t -44.0000\telaiss1_cat_f05")
    print("XMM-LSS\t\t 35.3333\t  -4.5000\txmm_cat_s05")
    print("Chandra-S\t 53.0000\t -28.2667\tchandra_cat_f05")
    print("Lockman\t\t161.2500\t +58.0000\tlockman_cat_s05")
    print("ELAIS-N1\t242.7500\t +55.0000\telaisn1_cat_s05")
    print("ELAIS-N2\t249.2000\t +41.0292\telaisn2_cat_s05")

def summon_files(folder_path):
    """
    For my convenience. Please do not use this as the specifics may not apply.
    """  
    descriptors = ['SpARCS_','.fits']
    files_SpA=make_file_list(folder_path, "SpARCS_IDs.txt", descriptors)
    descriptors = ['SVA1_COADD-','.fits']
    files_SVA=make_file_list(folder_path, "SVA1_IDs.txt", descriptors)
    return files_SpA+files_SVA

def plot_on_sky(folder_path,input_file,row_range,input_cols):
    """
    Plot sources from an excel spreadsheet with RA-DEC coords given in degrees
    onto an all-sky world map. Note: convert HMS and DMS coords into degrees 
    beforehand.
    """
    from openpyxl import load_workbook
    from astropy.coordinates import Angle
    import astropy.units as u
    import matplotlib.pyplot as plt
    wb = load_workbook(folder_path+input_file)
    sheet = wb.active
    ra,dec = [],[]
    for i in range(row_range[0],row_range[1]):
        ra.append(sheet[input_cols[0]+str(i)].value)
        dec.append(sheet[input_cols[1]+str(i)].value)   
    ra = Angle(ra*u.degree).wrap_at(180*u.degree)
    dec = Angle(dec*u.degree)
    fig = plt.figure(figsize=(8,6))
    ax = fig.add_subplot(111, projection="mollweide")
    ax.scatter(ra.radian, dec.radian, c='blue')
    ax.set_xticklabels(['14h','16h','18h','20h','22h','0h','2h','4h','6h','8h','10h'])
    ax.grid(True)
    fig.show()
    return

def make_ID_list(folder_path, output_file, identifier, ID_range):
    """
    Create an ID list for looping over a set of files with a specified 
    identifier and save into a text file for future reference.
    """
    from os import listdir
    output = open(folder_path+output_file,"a")
    content = listdir(folder_path)
    for item in content:
        if item[0:len(identifier)] == identifier:
            output.write(item[ID_range[0]:ID_range[1]]+"\n")
    output.close()
    return

def make_file_list(folder_path, filename, descriptors):
    """
    Build a list of file names given a text file storing IDs and descriptors
    to wrap around the IDs.
    """
    tf = open(folder_path+filename,"r")
    contents = tf.readlines()
    tf.close()
    ID_list=[]
    for line in contents:
        ID_list.append(descriptors[0]+(str(line)[:-1])+descriptors[1])
    return ID_list

def extract_reduced_headers(folder_path,input_file,output_file,descriptors,cols,flags):
    """
    Loop over a list of fits files and extract the requested header information 
    for each page of each fits file. Save the information into specified 
    columns of an excel spreadsheet.
    """
    from openpyxl import Workbook
    from astropy.io import fits
    tf = open(folder_path+input_file,"r")
    contents = tf.readlines()
    tf.close()
    ID_list=[]
    for line in contents:
        ID_list.append(str(line)[:-1])
    wb = Workbook()
    sheet = wb.active
    for i in range(len(cols)):
        sheet[cols[i]+str(1)]=flags[i]
    row_index = 1
    for ID in ID_list:
        a = fits.open(folder_path+descriptors[0]+ID+descriptors[1])
        for ext in range(len(a)):
            row_index = row_index+1
            for i in range(len(flags)):
                if flags[i] is "FITSFILE":
                    sheet[cols[i]+str(row_index)] = ID
                elif flags[i] is "INDEX":
                    sheet[cols[i]+str(row_index)] = str(ext)
                elif flags[i] in a[ext].header:
                    sheet[cols[i]+str(row_index)] = str(a[ext].header[flags[i]])
    wb.save(folder_path+output_file)
    
def unlock_database(folder_path,date_file,loc_file,ranges):
    """
    Build a list containing four dictionaries: the f->t dict, the f->c dict, 
    the t->f dict and the t->c dict, where f is the filename, t is the 
    timestamp and c is the coordinates. Requires two premade excel files with 
    a specific format which I will not bother to clarify in this comment. 
    Note: I have no scripts that generate these for you. Have fun!
    """  
    from openpyxl import load_workbook
    from collections import defaultdict
    
    wbd = load_workbook(folder_path+date_file)
    sheet_dates = wbd.active
    ft = defaultdict(list) #filename = [(datetime_start,datetime_end),...]
    id_name,date=-1,-1
    for i in range(ranges[0][0],ranges[0][1]):
        if sheet_dates['A'+str(i)].value:
            id_name = str(sheet_dates['A'+str(i)].value)+str(sheet_dates['B'+str(i)].value)
        if sheet_dates['C'+str(i)].value:
            date = str(sheet_dates['C'+str(i)].value)+":"
        time_start = date+str(sheet_dates['D'+str(i)].value)
        time_end = date+str(sheet_dates['E'+str(i)].value)
        ft[id_name].append((time_start,time_end))
        
    wbl = load_workbook(folder_path+loc_file)
    sheet_locs = wbl.active
    fc = defaultdict(list) #filename = (ra,dec)
    for i in range(ranges[1][0],ranges[1][1]):       
        id_name = str(sheet_locs['A'+str(i)].value)+str(sheet_locs['B'+str(i)].value)
        ra_dec = (sheet_locs['C'+str(i)].value,sheet_locs['D'+str(i)].value)
        fc[id_name].append(ra_dec)

    tf = defaultdict(list) #(datetime_start,datetime_end) = [filename,...]
    for key, value in ft.items():
        for v in value:
            tf[v].append(key)
        
    tc = defaultdict(list) #(datetime_start,datetime_end) = [(ra,dec),...]
    for key in tf:
        for value in tf[key]:
            tc[key].append(fc[value])
    
    return [ft,fc,tf,tc]

